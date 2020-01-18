'''
Change Log :

+ Removed some mis match code.
+ Refactor Analysis function
+ Sub Analysis appeared column => No. of pass + No. of fail


'''


'''
Change Required :

Total marks
Refactor result processing code



'''

from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.urls import NoReverseMatch, reverse
from os import system
from .models import year
from os import remove
import numpy as np
import mysql.connector
import pandas as pd
from django.contrib.auth.models import User, auth
from django.contrib import messages

def templetInstall(nameOfDepen):
    print('Installing '+ nameOfDepen )
    system('pip install ' + nameOfDepen)

# bs4 for scraping the result from web

# --------- Global Variables ------------------

'''
# for local use
user = 'root'
passwd = 'toor'
host = 'localhost'

'''
# for hosting
host = 'abhinavornikkoo.mysql.pythonanywhere-services.com'
user = 'abhinavornikkoo'
passwd = 'QWae@6Nsnshr59H'

try:
	from bs4 import BeautifulSoup as soup
except:
	templetInstall('bs4')
	from bs4 import BeautifulSoup as soup

# Lib for send request to Vtu server

try:
	import requests
except:
	templetInstall('requests')
	import requests

# Selecting the Hashed value in HTML scrapped page

try:
	import subprocess
except:
	templetInstall('subprocess')
	import subprocess

try:
	import re
except :
	templetInstall('re')
	import re

# Editing and Creating excel file 

try:
	import openpyxl
	from openpyxl.styles import PatternFill
except :
	templetInstall('openpyxl')
	import openpyxl
	from openpyxl.styles import PatternFill

# Connectivity to DBMS

from sqlalchemy.types import VARCHAR
import pymysql
import sqlalchemy
from sqlalchemy import create_engine


def captcha(butUrl,headers,cookies):
    captcha_url = butUrl + 'captcha_new.php'
    t = requests.get(captcha_url, headers=headers,cookies = cookies ,verify = False)
    with open("G://vtu_result//static//capt.png","wb") as f:
        f.write(t.content)
        f.close()


def recivecaptha(request):
	return render(request, 'captcha.html')

def getUrl(url):
	url = list(url)
	while(url[-1] != '/'):
		url = url[:-1]
	url = ''.join(url)
	return url


def colorFile(t,sub_analysis,topper, students_div, pass_due_to_ia):
	file_name = 'static/res.xlsx'
	with pd.ExcelWriter(file_name , engine="openpyxl", ) as writer:
		t.to_excel(writer, sheet_name = 'Result')
		sub_analysis.to_excel(writer, sheet_name = 'Analysis' , startrow = 2 , startcol = 2)
		students_div.to_excel(writer, sheet_name = 'Analysis' , startrow = len(students_div)+6 , startcol = 2)
		pass_due_to_ia.to_excel(writer, sheet_name = 'Analysis' , startrow = len(students_div)+len(sub_analysis)+8 , startcol = 2)
		topper.to_excel(writer, sheet_name = 'Analysis' , startrow = len(students_div) + len(sub_analysis) + len(pass_due_to_ia) + 12 , startcol = 2)
	wb = openpyxl.load_workbook(file_name)
	sheet = wb.active
	for row in range(4,sheet.max_row+1):
	    for column in range(6,sheet.max_column+1,4):
	        res = sheet.cell(row = row, column = column).value
	        if res == 'P':
	            pass
	        elif res == 'F':
	            sheet.cell(row = row, column = column-1).fill = PatternFill(start_color="FF0000", end_color="FFC7CE", fill_type = "solid")
	        elif res == 'A':
	            sheet.cell(row = row, column = column-1).fill = PatternFill(start_color="7FFFD4", end_color="FFC7CE", fill_type = "solid")
	        elif res == 'X':
	            sheet.cell(row = row, column = column-1).fill = PatternFill(start_color="ffff00", end_color="FFC7CE", fill_type = "solid")
	wb.save(file_name)

def colorFile_db(t,sub_analysis,topper , students_div, pass_due_to_ia):
	file_name = 'static/res.xlsx'
	with pd.ExcelWriter(file_name , engine="openpyxl", ) as writer:
		t.to_excel(writer, sheet_name = 'Result')
		sub_analysis.to_excel(writer, sheet_name = 'Analysis' , startrow = 2 , startcol = 2)
		students_div.to_excel(writer, sheet_name = 'Analysis' , startrow = len(students_div)+6 , startcol = 2)
		pass_due_to_ia.to_excel(writer, sheet_name = 'Analysis' , startrow = len(students_div)+len(sub_analysis)+8 , startcol = 2)
		topper.to_excel(writer, sheet_name = 'Analysis' , startrow = len(students_div) + len(sub_analysis) + len(pass_due_to_ia) + 12 , startcol = 2)
	wb = openpyxl.load_workbook(file_name)
	sheet = wb.active
	for row in range(4,sheet.max_row+1):
	    for column in range(5,sheet.max_column+1,4):
	        res = sheet.cell(row = row, column = column).value
	        if res == 'P':
	            pass
	        elif res == 'F':
	            sheet.cell(row = row, column = column+1).fill = PatternFill(start_color="FF0000", end_color="FFC7CE", fill_type = "solid")
	        elif res == 'A':
	            sheet.cell(row = row, column = column+1).fill = PatternFill(start_color="7FFFD4", end_color="FFC7CE", fill_type = "solid")
	        elif res == 'X':
	            sheet.cell(row = row, column = column+1).fill = PatternFill(start_color="ffff00", end_color="FFC7CE", fill_type = "solid")
	wb.save(file_name)


import time

# Create your views here.

def add(request):
	try:

		# rawUrl Stores the result age

		rawUrl = 'https://results.vtu.ac.in/vitavicbcsjj19/index.php'
		global butUrl
		butUrl = getUrl(rawUrl)
		print(butUrl)
		global link1
		link1 = butUrl + 'resultpage.php'

		# Header of a Browser 

		global headers
		headers = {'User-Agent': 'Mozilla/5.0'}

		# Saving the session to send request repeatedly 

		global session
		session = requests.Session()
		resp = session.post(link1, headers = headers, verify = False)
		global cookies
		print(requests)
		cookies = requests.utils.cookiejar_from_dict(requests.utils.dict_from_cookiejar(session.cookies))
		global csrf
		csrf = re.findall('[a-zA-Z0-9+/]{142}==', resp.text)
		captcha(butUrl,headers,cookies)
	except:
		redirect(home)

	return redirect(getcaptcha)


def getcaptcha(request ):
	if  request.user.is_authenticated:
		years = year.objects.all()
		return render(request, 'captcha.html', {'years' : years,
		'sem': range(1,9) })
	else:
		return redirect('home')

# Home Page request

def home(request):
		return render(request, 'home.html')

# About page request

def about(request):
		return render(request, 'about.html')

def Total(x, subject):
	z = 0
	for sub in subject:
		#print(z)
		z += x[sub]['Total']
	return z

def grade(x):
    if x.Result != 'P':
        return 
    if x.Total >= 70:
        return 'FCD'
    elif x.Total >= 60:
        return 'FC'
    else:
        return 'SC'


def results(request):

	# removing the Capthca image incase of cache file remain saved
	# use CTRL + F5 to clear cache

	#remove("G://vtu_result//static//capt.png")
	captain = str(request.POST['cap'])
	print(str(request.POST['sec']), str(request.POST['sem']) ,str(request.POST['year']))
	file = request.POST['csv']

	# Total marks is assume to be 800

	total_marks = 800
	wb1 = openpyxl.load_workbook(file)
	sheet1 = wb1.active
	l = []

	# Appending Usn excel sheet

	for x in range( 1, sheet1.max_row+1):
		l.append(sheet1.cell(column = 1 , row = x).value)
	
	data = {}
	data[("Name", '')] = {}
	flag = True

	# Sending request

	for usn in l:
		payload = {'lns': usn,
			'captchacode':captain,
			'token': csrf,
			'current_url': butUrl+'index.php'
			}
		resp = session.post(link1, headers=headers ,data = payload, cookies = cookies ,allow_redirects=False, verify = False)
		page_soup = soup(resp.text, 'html.parser')

		# Scraping the result table from page

		container = page_soup.findAll("div", {'class':'divTableRow'})
		fail = False

		# If the Usn is invalid or no result for a Usn

		if len(container) <= 1 :
			continue
		print(payload['lns'])
		nameConainter = page_soup.findAll('td', {'style':'padding-left:15px'})
		name  = nameConainter[0].text
		print(name[2:])
		Usn = payload['lns']
		data[("Name",'')][Usn] = name[2:]
		for con in range(1,len(container)-1):

			marks = container[con].findAll('div', {'class':'divTableCell'})

			### Per Subject Loop

			if marks[0].text == 'Subject Code':
					break
			print(marks[0].text ,marks[2].text+ ',' + marks[3].text + ',' +  marks[4].text + ',', marks[5].text )

			# Try catch if Key is not found in Dict

			try:
				data[(marks[0].text,"Internal")][Usn] = marks[2].text
			except(KeyError):
				data[(marks[0].text,"Internal")] = {}
				data[(marks[0].text,"Internal")][Usn] = marks[2].text
				
			# If External key not found create the External key else insert element in Dict

			try:
				data[(marks[0].text,"External")][Usn] = marks[3].text
			except(KeyError):
				data[(marks[0].text,"External")] = {}
				data[(marks[0].text,"External")][Usn] = marks[3].text

			# If Total key not found in Dict, create the result key and insert element in Dict

			try:
				data[(marks[0].text,"Total")][Usn] = marks[4].text
			except(KeyError):
				data[(marks[0].text,"Total")] = {}
				data[(marks[0].text,"Total")][Usn] = marks[4].text

			# If result key not found in Dict, create the result key and insert element in Dict

			try:
				data[(marks[0].text,"Result")][Usn] = marks[5].text
			except(KeyError):
				data[(marks[0].text,"Result")] = {}
				data[(marks[0].text,"Result")][Usn] = marks[5].text
	
	# Casting result into Dataframe

	t = pd.DataFrame(data)

	# If row size of Dataframe is 0 ie no data is fetched

	if t.shape[0] == 0:
		messages.error(request,'Invalid Capthcha ! If problem presist try clearing Cache')
		return redirect('fetch_result')

	z = t.set_index('Name' , append= True).stack(0)
	z.index.names = ['Usn','Name', 'Subject']
	z['sec'] = str(request.POST['sec'])
	sem = str(request.POST['sem'])
	mydb = connect_db()
	conn = mydb.cursor()
	database = str(request.POST['year'])
	conn.execute("CREATE DATABASE IF NOT EXISTS `{0}` ".format(database))
	conn.close()
	conn = create_engine('mysql+pymysql://' + user + ':' + passwd + '@' + host + '/' + database , echo=False)

	# Saving the Dataframe in Database

	z.to_sql(name=sem, con=conn, if_exists = 'append' ,dtype={'Usn': VARCHAR(10),'Name': VARCHAR(50), 'Subject': VARCHAR(10)}) 
	#
	t.index.name = "Usn"
	subject=list(t.columns.levels[0][:9])
	for sub in subject:
		t.loc[:,(sub,['Internal','External','Total'])] = t.loc[:,(sub,['Internal','External','Total'])].astype('float64').astype('Int64')
	t['Grand Total'] = t.iloc[:,1:].fillna(0).apply(Total,args = [subject],axis = 1)
	detail = {}
	for sub in subject:
		detail[sub] = {}
	for sub in subject:
		df = t[sub]['Result'].value_counts()
		detail[sub] = { 'P': 0,
					'A': 0,
					'F': 0,
					}
		for x in df.index:
			detail[sub][x] = df[x]
		detail[sub]['appeared'] = t[sub]['Result'].count() 
	for x in detail.keys():
		for fc in ['FCD','FC','SC']:
			detail[x][fc] = 0
	for sub in subject:
		df = t[sub].apply(grade,axis = 1).value_counts()
		for x in df.index:
				detail[sub][x] = df[x]
	sub_analysis = pd.DataFrame.from_dict(detail,orient = 'index')
	sub_analysis['pass_%'] = sub_analysis['P']/sub_analysis['appeared']*100
	sub_analysis['pass_%'] = sub_analysis['pass_%'].apply(lambda x : round(x,2))
	sub_analysis['Avg Marks(EXT)'] = 0
	sub_analysis['Avg Marks(INT)'] = 0
	for sub in subject:
		sub_analysis.loc[sub ,'Avg Marks(EXT)'] = round(t[sub]['External'].mean(),2)
		sub_analysis.loc[sub ,'Avg Marks(INT)'] = round(t[sub]['Internal'].mean(),2)

	topper = t[['Name','Grand Total']].reset_index().sort_values('Grand Total', ascending = False).head().stack()
	topper.set_index('Usn',inplace = True)
	topper = topper[['Name','Grand Total' ]]
	pass_ia = {}

	# Passing marks for ia
	fail_ia_marks = 19
	for no in range(len(subject)+1):
		pass_ia['Total Students failed in '+str(no)] = {'Ext' : 0,
					'Int': 0}
	def pass_fail(x):
		ia, ext = 0,0  
		for sub in subject:
			if x[sub,'Result'] == 'F':
				if x[sub,'Internal'] < fail_ia_marks:
					ia += 1
				else:
					ext += 1
		pass_ia['Total Students failed in '+str(ia + ext)]['Ext'] += ext
		pass_ia['Total Students failed in '+str(ia + ext)]['Int'] += ia
	t.apply(pass_fail, axis = 1)
	del pass_ia['Total Students failed in ' + '0']
	pass_due_to_ia = pd.DataFrame.from_dict(pass_ia, orient = 'index')

	# Creating dic_class for student detail in a Class

	dic_class = {
		"STUDENT APPEARED" : 0,
		"FIRST CLASS DISTIN":  0,
		"FIRST CLASS" : 0,
		"SECOUND CLASS" : 0,
		"PASSED" : 0,
		"FAILED" : 0,
		"ABSENT": 0,
	}

	def student_grade(x,marks):
		if x*100/marks >= 70:
			return 'FIRST CLASS DISTIN'
		elif x*100/marks >= 60:
			return 'FIRST CLASS'
		else:
			return 'SECOUND CLASS'

	def dic_pass(x):
		dic_class['STUDENT APPEARED'] += 1
		absent, fail = 0, 0
		for sub in subject:
			if x[sub, 'Result'] == 'F':
				fail += 1
			elif x[sub, 'Result'] == 'A':
				absent += 1
		if fail == 0 and absent == 0:
			dic_class['PASSED'] += 1
			dic_class[student_grade(x['Grand Total'][0], total_marks)] += 1
			
		else:
			dic_class['FAILED'] += 1
		if absent :
			dic_class['ABSENT'] += 1

	temp = t.apply(dic_pass, axis = 1)

	# Changing the row 

	students_div = pd.DataFrame.from_dict(dic_class, orient = 'index', columns =['Freq'])
	students_div.loc["Overall Pass"] = int(students_div.loc['PASSED']*100/students_div.loc['STUDENT APPEARED'])

	colorFile(t,sub_analysis, topper, students_div, pass_due_to_ia)


	return render(request,'result.html',{
		'result': t.to_html(classes = 'table'),
		'sub_analysis':sub_analysis.to_html(classes = 'table'),
		'topper' : topper.to_html(classes = 'table'),
		'students_div' : students_div.to_html(classes = 'table'),
		'pass_due_to_ia' : pass_due_to_ia.to_html(classes = 'table')
		})

def result_db(request):
	return render(request,'result_db.html')

# Connecting to database

def connect_db():
	mydb = mysql.connector.connect(
		host = host,
		user = user,
		passwd = passwd,
	)
	return mydb


# Fetching results from Database

def analyze(t, total_marks = 800, fail_ia_marks = 19):
	t.index.name = "Usn"

	# Extracting the subjects from dataframe

	subject=list(t.columns.levels[0][:9])

	# Converting the datatype of result(Internal, External, Total) into int64
	# Converting nan to float64 and then to int64
	# This may change due to upcoming version ( now int64 support nan values )

	for sub in subject:
		t.loc[:,(sub,['Internal','External','Total'])] = t.loc[:,(sub,['Internal','External','Total'])].astype('float64').astype('Int64')
	
	# Filling the nan value into zero ( to avoid any error during the addition)
	# Creating "Grand Total" column in Dataframe for Total marks
	# args paramater in apply was tricky !!!!

	t['Grand Total'] = t.iloc[:,1:].fillna(0).apply(Total,args = [subject],axis = 1)


	detail = {}
	for sub in subject:
		detail[sub] = {}

	for sub in subject:

		# Getting pass fail student value count from dataframe of subjects

		df = t[sub]['Result'].value_counts()
		detail[sub] = { 'P': 0,
					'A': 0,
					'F': 0,
					}

		# Mapping the pass fail values into dataframe from count values

		for x in df.index:
			detail[sub][x] = df[x]

		# Student appeared is Student pass +  fail

		# detail[sub]['Appeared'] = t[sub]['Result'].count() -> Pervious version
		detail[sub]['Appeared'] = detail[sub]['P'] + detail[sub]['F']

	for x in detail.keys():
		for fc in ['FCD','FC','SC']:
			detail[x][fc] = 0
	
	
	for sub in subject:

		# assigning FCD and FC .. based on Total marks
		# And Mapping the Value count to sub_analysis dataframe
		df = t[sub].apply(grade,axis = 1).value_counts()
		for x in df.index:
				detail[sub][x] = df[x]
	sub_analysis = pd.DataFrame.from_dict(detail,orient = 'index')

	# pass_% = total student pass divided by total student appeared

	sub_analysis['Pass Percentage'] = sub_analysis['P']/sub_analysis['Appeared']*100

	# Rounding off the values upto 2 decimal palace

	sub_analysis['Pass Percentage'] = sub_analysis['Pass Percentage'].apply(lambda x : round(x,2))

	sub_analysis['Avg Marks(EXT)'] = 0
	sub_analysis['Avg Marks(INT)'] = 0
	for sub in subject:

		# Averge marks for internal and external are means of respective column
		# Mean does'nt include nan values either in numerator(Total value excluding the nan) or denominator(Total count excluding the nan)

		sub_analysis.loc[sub ,'Avg Marks(EXT)'] = round(t[sub]['External'].mean(),2)
		sub_analysis.loc[sub ,'Avg Marks(INT)'] = round(t[sub]['Internal'].mean(),2)


	# top 5 student based on Total marks

	topper = t[['Name','Grand Total']].reset_index().sort_values('Grand Total', ascending = False).head().stack()
	topper.set_index('Usn',inplace = True)
	topper = topper[['Name','Grand Total' ]]
	
	# Pass fail due to IA exam
	
	pass_ia = {}

	#
	for no in range(len(subject)+1):
		pass_ia['Total Students failed in '+str(no)] = {'Ext' : 0,
					'Int': 0}
	
	# If student fail then check Ia marks, If ia marks > fail_ia_marks else fail in External
	
	def pass_fail(x):
		ia, ext = 0,0  
		for sub in subject:
			if x[sub,'Result'] == 'F':

				# fail_ia_marks variable value from call function else default value is 19

				if x[sub,'Internal'] < fail_ia_marks:
					ia += 1
				else:
					ext += 1
		pass_ia['Total Students failed in '+str(ia + ext)]['Ext'] += ext
		pass_ia['Total Students failed in '+str(ia + ext)]['Int'] += ia
	t.apply(pass_fail, axis = 1)

	# remove Student fail in 0 subject

	del pass_ia['Total Students failed in ' + '0']
	pass_due_to_ia = pd.DataFrame.from_dict(pass_ia, orient = 'index')

	# Creating dic_class for student detail in a Class

	dic_class = {
		"STUDENT APPEARED" : 0,
		"FIRST CLASS DISTIN":  0,
		"FIRST CLASS" : 0,
		"SECOUND CLASS" : 0,
		"PASSED" : 0,
		"FAILED" : 0,
		"ABSENT": 0,
	}

	# FCD, FC, SC

	def student_grade(x,marks):
		if x*100/marks >= 70:
			return 'FIRST CLASS DISTIN'
		elif x*100/marks >= 60:
			return 'FIRST CLASS'
		else:
			return 'SECOUND CLASS'

	def dic_pass(x):
		dic_class['STUDENT APPEARED'] += 1
		absent, fail = 0, 0
		for sub in subject:
			if x[sub, 'Result'] == 'F':
				fail += 1
			elif x[sub, 'Result'] == 'A':
				absent += 1
		if fail == 0 and absent == 0:
			dic_class['PASSED'] += 1
			dic_class[student_grade(x['Grand Total'][0], total_marks)] += 1
			
		else:
			dic_class['FAILED'] += 1
		if absent :
			dic_class['ABSENT'] += 1

	# Using the row as Index for Dataframe
	t.apply(dic_pass, axis = 1)
	students_div = pd.DataFrame.from_dict(dic_class, orient = 'index', columns =['Freq'])
	students_div.loc["Overall Pass"] = int(students_div.loc['PASSED']*100/students_div.loc['STUDENT APPEARED'])

	return (t, sub_analysis, topper, students_div, pass_due_to_ia )


# Analysis Page 

def analysis_page(request):
	if request.method == 'POST':
		database = str(request.POST['year'])
		sem = str(request.POST['sem'])
		sec = str(request.POST['sec'])
		pass_ia_marks = int(request.POST['ia_marks'])
		total_marks = int(request.POST['total_marks'])
		conn = create_engine('mysql+pymysql://' + user + ':' + passwd + '@' + host + '/' + database , echo=False)
		try:
			t = pd.read_sql('SELECT * FROM' + ' `' + sem +'` ' + 'where sec = ' + '"' + sec + '"' , conn)
		except sqlalchemy.exc.ProgrammingError:

			# Redirect if Result is not found

			messages.error(request,'No result found for ' + sem + ' semister')
			return redirect ('analysis_page')

			# No. of row = 0, then no result found

		if t.shape[0] == 0:
			messages.error(request,'No result found for section ' +sec + ' of ' + sem + ' semister')
			return redirect ('analysis_page')

		
		# Re indexing the dataframe in result format

		t = t.set_index(["Usn","Name",'Subject']).drop('sec',axis = 1).sort_index().unstack().stack(0).unstack()
		t = t.reset_index().set_index("Usn")
		t, sub_analysis, topper, students_div, pass_due_to_ia = analyze(t, total_marks, pass_ia_marks)
		
		
		# Fill color in excel sheet 

		colorFile_db(t,sub_analysis, topper , students_div, pass_due_to_ia)

		return render(request,'result.html',{
			'result': t.to_html(classes = 'table'),
			'sub_analysis':sub_analysis.to_html(classes = 'table'),
			'topper' : topper.to_html(classes = 'table'),
			'students_div' : students_div.to_html(classes = 'table'),
			'pass_due_to_ia' : pass_due_to_ia.to_html(classes = 'table')
			})

	else:
		mydb = connect_db()
		conn = mydb.cursor()
		conn.execute("SHOW DATABASES")
		db_name = []
		for name in conn:
			if len(name[0]) == 4 and name[0].isnumeric() == True:
				db_name += list(name)
		conn.close()
		return render(request,'analysis.html' , {'db_name' : db_name})



def login(request):
	if request.method == 'POST':
		username = request.POST['username']
		password = request.POST['passwd']

		user = auth.authenticate(username = username, password = password)

		if user is not None:
			auth.login(request, user)
			messages.success(request,'Welcome ' + user.first_name)
			return redirect('home')
		else:
			messages.error(request, 'Invalid Ceredential')
			return redirect ('login')

	else:
		return render(request, 'login.html')

# Logout 

def logout(request):
	auth.logout(request)
	messages.success(request,'You have successfully logged out.')
	return redirect( 'home')

def student(request):
	if request.method == 'POST':
		pass
	else :
		return render(request, 'student.html')